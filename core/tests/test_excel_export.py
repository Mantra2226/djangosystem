"""
TEST SUITE: OpenPyXL Excel Export Engine & Dual-Stream Executive Reporting
(core/tests/test_excel_export.py)

Validates:
1. Admin changelist Excel exports (selected and full filtered datasets)
2. Number formatting masks (currency, decimal, dates)
3. Dual-stream executive dashboard multi-sheet exports
4. Complete preservation of existing PDF and CSV export actions
5. Mathematical reconciliation of COGS between summary and dispatch itemization
"""

from decimal import Decimal
import io
from datetime import date, timedelta

from django.test import TestCase, Client, RequestFactory
from django.urls import reverse
from django.utils import timezone
from django.contrib.auth.models import User
from django.contrib.admin.sites import site

import openpyxl

from core.models import (
    Supplier, Product, Inventory, StockTransaction,
    WorkOrder, MaterialVarianceRecord, SalesInvoice,
    SalesOrder, SalesOrderItem, Customer, DispatchRecord,
    PurchaseOrder, PurchaseOrderItem, FinanceEntry
)
from core.services.excel_export_service import NUMBER_FORMAT_CURRENCY
from core.admin import (
    InventoryAdmin, StockTransactionAdmin, WorkOrderAdmin,
    MaterialVarianceRecordAdmin, PurchaseOrderAdmin, SalesInvoiceAdmin,
    FinanceEntryAdmin, export_as_csv
)


class ExcelExportEngineTestCase(TestCase):
    def setUp(self):
        self.client = Client()
        self.factory = RequestFactory()

        # Create staff user for admin and report access
        self.staff_user = User.objects.create_superuser(
            username='admin_excel',
            password='password123',
            email='admin@example.com'
        )
        self.client.login(username='admin_excel', password='password123')

        # Baseline entities
        self.supplier = Supplier.objects.create(name='Acme Chemical Corp', contact_info='acme@test.com')
        self.customer = Customer.objects.create(
            customer_name='Premier Glazing Ltd',
            contact_info='info@premier.test',
            shipping_address='123 Industrial Way'
        )

        # Products
        self.raw_mat = Product.objects.create(
            name='Whiting Powder',
            sku='RAW-WHT-01',
            product_type='RAW',
            category='Powder',
            unit_of_measurement='kg',
            supplier=self.supplier
        )
        self.pkg_mat = Product.objects.create(
            name='5L Metal Can',
            sku='PKG-CAN-01',
            product_type='RAW',
            category='Packaging',
            unit_of_measurement='pcs',
            supplier=self.supplier
        )
        self.pigment = Product.objects.create(
            name='Black Carbon Pigment',
            sku='RAW-BLK-01',
            product_type='RAW',
            category='Pigment',
            unit_of_measurement='kg',
            supplier=self.supplier
        )
        self.finished_prod = Product.objects.create(
            name='Linseed Glazing Putty 5kg',
            sku='FG-PUT-05',
            product_type='FINISHED',
            category='Putty',
            unit_of_measurement='tub',
            selling_price=Decimal('25.00')
        )

        # Inventory
        self.inv_raw = Inventory.objects.create(
            product=self.raw_mat,
            quantity_available=Decimal('500.00'),
            quantity_allocated=Decimal('50.00'),
            unit_cost=Decimal('2.50'),
            location='Warehouse Section A'
        )
        self.inv_pkg = Inventory.objects.create(
            product=self.pkg_mat,
            quantity_available=Decimal('5.00'),  # Low stock (<= 10)
            quantity_allocated=Decimal('0.00'),
            unit_cost=Decimal('1.80'),
            location='Warehouse Section B'
        )
        self.inv_pigment = Inventory.objects.create(
            product=self.pigment,
            quantity_available=Decimal('3.00'),  # Low stock (<= 10)
            quantity_allocated=Decimal('0.00'),
            unit_cost=Decimal('4.50'),
            location='Warehouse Section C'
        )
        self.inv_fg = Inventory.objects.filter(product=self.finished_prod).first()
        self.inv_fg.quantity_available = Decimal('50.00')
        self.inv_fg.unit_cost = Decimal('12.00')
        self.inv_fg.location = 'Finished Goods Bay'
        self.inv_fg.save()

        # Work Order
        self.work_order = WorkOrder.objects.create(
            work_order_code='WO-TEST-001',
            product=self.finished_prod,
            quantity_produced=Decimal('100.00'),
            actual_quantity_produced=Decimal('98.00'),
            scrap_quantity=Decimal('2.00'),
            status='COMPLETED',
            production_start_date=timezone.now().date()
        )

        # Material Variance
        self.variance = MaterialVarianceRecord.objects.create(
            work_order=self.work_order,
            product=self.raw_mat,
            quantity_expected=Decimal('100.00'),
            quantity_actual=Decimal('105.00'),
            quantity_variance=Decimal('5.00'),
            unit_cost=Decimal('2.50'),
            financial_impact=Decimal('12.50'),
            variance_classification='UNFAVOURABLE'
        )

        # Sales Order & Dispatch
        self.sales_order = SalesOrder.objects.create(
            customer=self.customer,
            status='approved'
        )
        self.so_item = SalesOrderItem.objects.create(
            sales_order=self.sales_order,
            product=self.finished_prod,
            quantity_ordered=Decimal('20.00'),
            unit_price=Decimal('25.00')
        )
        self.dispatch = DispatchRecord.objects.create(
            customer=self.customer,
            sales_order_item=self.so_item,
            product=self.finished_prod,
            quantity_dispatched=Decimal('20.00'),
            status='delivered',
            dispatch_date=timezone.now().date()
        )

        # Sales Invoice
        self.invoice = SalesInvoice.objects.create(
            customer=self.customer,
            sales_order=self.sales_order,
            dispatch=self.dispatch,
            invoice_date=timezone.now().date(),
            subtotal=Decimal('500.00'),
            tax_amount=Decimal('80.00'),
            total_amount=Decimal('580.00'),
            status='POSTED'
        )

        # Purchase Order
        self.po = PurchaseOrder.objects.create(
            supplier=self.supplier,
            status='RECEIVED',
            order_date=timezone.now().date()
        )
        self.po_item = PurchaseOrderItem.objects.create(
            purchase_order=self.po,
            product=self.raw_mat,
            quantity_ordered=Decimal('200.00'),
            price_per_unit=Decimal('2.50')
        )

    def test_admin_excel_export_preserves_filters(self):
        """
        Validates that changelist export respects filtered querysets.
        """
        admin_instance = InventoryAdmin(Inventory, site)
        request = self.factory.get('/admin/core/inventory/?low_stock=true')
        request.user = self.staff_user

        # Filtered queryset for low stock items (only PKG and FG have <= 10 units)
        cl = admin_instance.get_changelist_instance(request)
        filtered_qs = cl.get_queryset(request)

        self.assertEqual(filtered_qs.count(), 2)

        # Trigger top-bar export action
        response = admin_instance.action_export_all_to_excel(request)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        # Load workbook and check row count (1 header + 2 data rows = 3 rows)
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        self.assertEqual(ws.max_row, 3)

    def test_numeric_and_currency_formatting(self):
        """
        Validates cell formatting masks: currency ($#,##0.00), decimal (#,##0.00), date (yyyy-mm-dd).
        """
        admin_instance = SalesInvoiceAdmin(SalesInvoice, site)
        request = self.factory.get('/admin/core/salesinvoice/')
        request.user = self.staff_user

        response = admin_instance.action_export_all_to_excel(request)
        self.assertEqual(response.status_code, 200)

        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active

        # Find column indices for Subtotal and Invoice Date
        subtotal_col = None
        date_col = None
        for col_idx in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col_idx).value
            if header == 'Subtotal':
                subtotal_col = col_idx
            elif header == 'Invoice Date':
                date_col = col_idx

        self.assertIsNotNone(subtotal_col)
        self.assertIsNotNone(date_col)

        # Check format on data row (row 2)
        subtotal_cell = ws.cell(row=2, column=subtotal_col)
        self.assertEqual(subtotal_cell.number_format, NUMBER_FORMAT_CURRENCY)
        self.assertEqual(float(subtotal_cell.value), 500.00)

        date_cell = ws.cell(row=2, column=date_col)
        self.assertEqual(date_cell.number_format, 'yyyy-mm-dd')

    def test_dual_dashboard_exports_contain_correct_sheets(self):
        """
        Validates dual-stream dashboard export views:
        - Financial Analytics: 3 sheets (P&L Summary, Sales Invoices, COGS Dispatches)
        - Shop-Floor Analytics: 3 sheets (Completed Builds, Material Variances, Low-Stock Alerts)
        """
        # 1. Financial Analytics Export
        fin_url = reverse('export_financial_analytics_excel')
        fin_resp = self.client.get(fin_url)
        self.assertEqual(fin_resp.status_code, 200)
        self.assertEqual(fin_resp['Content-Type'], 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        fin_wb = openpyxl.load_workbook(io.BytesIO(fin_resp.content))
        self.assertEqual(fin_wb.sheetnames, ['P&L Summary', 'Sales Invoices', 'COGS Dispatches'])

        # 2. Shop-Floor Analytics Export
        sf_url = reverse('export_shopfloor_analytics_excel')
        sf_resp = self.client.get(sf_url)
        self.assertEqual(sf_resp.status_code, 200)
        self.assertEqual(sf_resp['Content-Type'], 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        sf_wb = openpyxl.load_workbook(io.BytesIO(sf_resp.content))
        self.assertEqual(sf_wb.sheetnames, ['Completed Builds', 'Material Variances', 'Low-Stock Alerts'])

    def test_cogs_reconciliation_between_summary_and_dispatches(self):
        """
        Validates that headline COGS on Sheet 1 mathematically reconciles with
        itemized dispatch lines on Sheet 3.
        """
        fin_url = reverse('export_financial_analytics_excel')
        response = self.client.get(fin_url)
        wb = openpyxl.load_workbook(io.BytesIO(response.content))

        # Sheet 1: Extract COGS
        ws_pnl = wb['P&L Summary']
        headline_cogs = None
        for row_idx in range(2, ws_pnl.max_row + 1):
            line_name = ws_pnl.cell(row=row_idx, column=1).value
            if line_name == 'Cost of Goods Sold (COGS)':
                headline_cogs = Decimal(str(ws_pnl.cell(row=row_idx, column=2).value))
                break

        self.assertIsNotNone(headline_cogs)

        # Sheet 3: Sum dispatch lines
        ws_disp = wb['COGS Dispatches']
        cogs_col = None
        for col_idx in range(1, ws_disp.max_column + 1):
            if ws_disp.cell(row=1, column=col_idx).value in ('COGS Line Valuation (KSh)', 'COGS Line Valuation'):
                cogs_col = col_idx
                break

        self.assertIsNotNone(cogs_col)

        itemized_cogs_sum = Decimal('0.00')
        for row_idx in range(2, ws_disp.max_row + 1):
            val = ws_disp.cell(row=row_idx, column=cogs_col).value
            if val is not None:
                itemized_cogs_sum += Decimal(str(val))

        self.assertEqual(headline_cogs, itemized_cogs_sum)

    def test_pdf_and_csv_export_actions_unaffected(self):
        """
        Assures that adding Excel exports introduced zero regressions:
        - export_as_csv remains active in admin actions
        - action_download_pdf remains in actions_detail
        - PDF generation still returns valid PDF MIME responses
        """
        # CSV action checks
        self.assertIn(export_as_csv, InventoryAdmin.actions)
        self.assertIn(export_as_csv, StockTransactionAdmin.actions)
        self.assertIn(export_as_csv, WorkOrderAdmin.actions)
        self.assertIn(export_as_csv, MaterialVarianceRecordAdmin.actions)
        self.assertIn(export_as_csv, PurchaseOrderAdmin.actions)
        self.assertIn(export_as_csv, SalesInvoiceAdmin.actions)

        # Detail PDF action checks
        self.assertIn('action_download_pdf', SalesInvoiceAdmin.actions_detail)
        self.assertIn('action_download_pdf', FinanceEntryAdmin.actions_detail)

        # Execute PDF download action directly
        invoice_admin = SalesInvoiceAdmin(SalesInvoice, site)
        pdf_resp = invoice_admin.action_download_pdf(None, self.invoice.pk)
        self.assertEqual(pdf_resp.status_code, 200)
        self.assertEqual(pdf_resp['Content-Type'], 'application/pdf')
        self.assertTrue(pdf_resp.content.startswith(b'%PDF'))

    def test_financial_export_preview_json(self):
        """
        Validates that GET /analytics/export/financial/?preview=1 returns
        a JSON payload with 3 serialized sheets for modal spreadsheet previewing.
        """
        url = reverse('export_financial_analytics_excel') + '?preview=1'
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/json')

        data = response.json()
        self.assertEqual(data['stream'], 'financial')
        self.assertEqual(len(data['sheets']), 3)

        sheet_titles = [s['title'] for s in data['sheets']]
        self.assertEqual(sheet_titles, ['P&L Summary', 'Sales Invoices', 'COGS Dispatches'])

        # Verify P&L Summary sheet has line items and formatting metadata
        pnl_sheet = data['sheets'][0]
        self.assertGreater(pnl_sheet['row_count'], 0)
        self.assertEqual(pnl_sheet['headers'], ['Financial Statement Line', 'Amount (KSh)', 'Classification / Notes'])
        self.assertEqual(pnl_sheet['formats'], ['text', 'currency', 'text'])

    def test_shopfloor_export_preview_json(self):
        """
        Validates that GET /analytics/export/shopfloor/?preview=1 returns
        a JSON payload with 3 serialized sheets for modal spreadsheet previewing.
        """
        url = reverse('export_shopfloor_analytics_excel') + '?preview=1'
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/json')

        data = response.json()
        self.assertEqual(data['stream'], 'shopfloor')
        self.assertEqual(len(data['sheets']), 3)

        sheet_titles = [s['title'] for s in data['sheets']]
        self.assertEqual(sheet_titles, ['Completed Builds', 'Material Variances', 'Low-Stock Alerts'])

        # Verify Low-Stock Alerts sheet structure
        alert_sheet = data['sheets'][2]
        self.assertIn('Product Name', alert_sheet['headers'])
        self.assertIn('Stock Deficit', alert_sheet['headers'])
        self.assertGreaterEqual(alert_sheet['row_count'], 1)

    def test_dashboard_page_renders_preview_buttons_and_modal(self):
        """
        Validates that the reports dashboard template renders the export preview buttons
        with openExcelPreview handlers and includes the spreadsheet modal markup.
        """
        url = reverse('reports_dashboard')
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)

        content = response.content.decode('utf-8')
        self.assertIn("openExcelPreview('financial')", content)
        self.assertIn("openExcelPreview('shopfloor')", content)
        self.assertIn('id="excelPreviewModal"', content)
        self.assertIn('id="modalDownloadBtn"', content)
