"""
EXCEL EXPORT SERVICE (core/services/excel_export_service.py)

Provides corporate-styled OpenPyXL workbook generation for single-sheet
admin changelist querysets and multi-sheet executive dashboard analytics.
"""

from decimal import Decimal
from datetime import date, datetime
import io

from django.http import HttpResponse
from django.utils import timezone
from django.db.models import QuerySet

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# Corporate styling constants
NAVY_HEADER_FILL = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
HEADER_FONT = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
DATA_FONT = Font(name='Calibri', size=10, color='0F172A')

THIN_BORDER_COLOR = 'CBD5E1'
CELL_BORDER = Border(
    left=Side(style='thin', color=THIN_BORDER_COLOR),
    right=Side(style='thin', color=THIN_BORDER_COLOR),
    top=Side(style='thin', color=THIN_BORDER_COLOR),
    bottom=Side(style='thin', color=THIN_BORDER_COLOR)
)

ALIGN_LEFT = Alignment(horizontal='left', vertical='center')
ALIGN_RIGHT = Alignment(horizontal='right', vertical='center')
ALIGN_CENTER = Alignment(horizontal='center', vertical='center')

from django.conf import settings

CURRENCY_SYMBOL = getattr(settings, 'CURRENCY_SYMBOL', 'KSh')
NUMBER_FORMAT_CURRENCY = f'"{CURRENCY_SYMBOL} " #,##0.00;[Red]-"{CURRENCY_SYMBOL} " #,##0.00;"{CURRENCY_SYMBOL} " 0.00'
NUMBER_FORMAT_DECIMAL = '#,##0.00'
NUMBER_FORMAT_INTEGER = '#,##0'
NUMBER_FORMAT_DATE = 'yyyy-mm-dd'
NUMBER_FORMAT_DATETIME = 'yyyy-mm-dd hh:mm'


def resolve_attribute(obj, attr_or_callable):
    """
    Safely resolves attribute paths on an object, supporting:
    - Direct attributes: 'status'
    - Dotted paths with null foreign-key guards: 'product.name', 'supplier.supplier_code'
    - Callables / model methods: 'get_status_display'
    - Standalone callable functions taking the object as first argument
    """
    if obj is None:
        return None

    if callable(attr_or_callable):
        try:
            return attr_or_callable(obj)
        except Exception:
            return None

    current = obj
    for part in str(attr_or_callable).split('.'):
        if current is None:
            return None
        if hasattr(current, part):
            val = getattr(current, part)
            if callable(val) and not isinstance(val, type):
                try:
                    current = val()
                except Exception:
                    current = None
            else:
                current = val
        else:
            return None
    return current


def format_cell_value(cell, value, format_type):
    """
    Applies explicit values, typography, borders, and number format masks to a cell.
    """
    cell.font = DATA_FONT
    cell.border = CELL_BORDER

    if value is None:
        cell.value = ''
        cell.alignment = ALIGN_LEFT
        return

    fmt = (format_type or 'text').lower()

    if fmt == 'currency':
        try:
            val = float(value) if isinstance(value, (Decimal, int, float)) else float(str(value).replace('$', '').replace(CURRENCY_SYMBOL, '').replace(',', '').strip())
            cell.value = val
        except (ValueError, TypeError):
            cell.value = str(value)
        cell.number_format = NUMBER_FORMAT_CURRENCY
        cell.alignment = ALIGN_RIGHT

    elif fmt in ('decimal', 'quantity', 'float'):
        try:
            val = float(value) if isinstance(value, (Decimal, int, float)) else float(str(value).replace(',', ''))
            cell.value = val
        except (ValueError, TypeError):
            cell.value = str(value)
        cell.number_format = NUMBER_FORMAT_DECIMAL
        cell.alignment = ALIGN_RIGHT

    elif fmt in ('integer', 'int'):
        try:
            val = int(value) if isinstance(value, (int, float, Decimal)) else int(str(value).replace(',', ''))
            cell.value = val
        except (ValueError, TypeError):
            cell.value = str(value)
        cell.number_format = NUMBER_FORMAT_INTEGER
        cell.alignment = ALIGN_RIGHT

    elif fmt == 'date':
        if isinstance(value, (datetime, date)):
            if isinstance(value, datetime):
                value = value.date()
            cell.value = value.strftime('%Y-%m-%d')
        else:
            cell.value = str(value)
        cell.number_format = NUMBER_FORMAT_DATE
        cell.alignment = ALIGN_CENTER

    elif fmt == 'datetime':
        if isinstance(value, datetime):
            cell.value = value.strftime('%Y-%m-%d %H:%M')
        elif isinstance(value, date):
            cell.value = value.strftime('%Y-%m-%d')
        else:
            cell.value = str(value)
        cell.number_format = NUMBER_FORMAT_DATETIME
        cell.alignment = ALIGN_CENTER

    else:
        cell.value = str(value)
        cell.alignment = ALIGN_LEFT


def apply_worksheet_styling(ws, freeze_header=True):
    """
    Applies standard corporate styling across an openpyxl worksheet:
    - #1E293B Dark Navy fill with bold white text on Header (Row 1)
    - Row heights: 26pt for header, 20pt for data rows
    - Enables gridlines explicitly
    - Freezes the top header row at A2
    - Auto-computes column widths: max(length + 4, 12)
    """
    # Explicitly enable gridlines
    ws.sheet_view.showGridLines = True

    # Header row height
    ws.row_dimensions[1].height = 26.0

    # Apply header styles to Row 1
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = NAVY_HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = ALIGN_CENTER
        cell.border = CELL_BORDER

    # Set data row heights and compute column widths
    col_max_lengths = {col_idx: len(str(ws.cell(row=1, column=col_idx).value or '')) for col_idx in range(1, ws.max_column + 1)}

    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 20.0
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            val_str = str(cell.value or '')
            if len(val_str) > col_max_lengths[col_idx]:
                col_max_lengths[col_idx] = len(val_str)

    # Set auto-computed column widths
    for col_idx, max_len in col_max_lengths.items():
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # Freeze header pane
    if freeze_header:
        ws.freeze_panes = 'A2'


def export_queryset_to_excel(queryset, fields_map, filename_prefix='export'):
    """
    Generates a single-sheet corporate Excel workbook from a Django QuerySet or list of objects.

    :param queryset: Django QuerySet or list of model instances
    :param fields_map: List of tuples: (attribute_or_callable, column_header, format_type)
                       format_type: 'currency' | 'decimal' | 'integer' | 'date' | 'datetime' | 'text'
    :param filename_prefix: Base string for the downloaded file name
    :return: HttpResponse with .xlsx content type
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = filename_prefix.replace('_', ' ').title()[:31]

    # Write headers
    headers = [col_def[1] for col_def in fields_map]
    ws.append(headers)

    # Write data rows
    for obj in queryset:
        row_cells = []
        row_num = ws.max_row + 1
        for col_idx, col_def in enumerate(fields_map, start=1):
            attr_path = col_def[0]
            fmt_type = col_def[2] if len(col_def) > 2 else 'text'
            val = resolve_attribute(obj, attr_path)
            cell = ws.cell(row=row_num, column=col_idx)
            format_cell_value(cell, val, fmt_type)

    apply_worksheet_styling(ws, freeze_header=True)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    date_str = timezone.now().strftime('%Y-%m-%d')
    filename = f"{filename_prefix}_{date_str}.xlsx"

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def build_multi_sheet_workbook(sheets_config, filename_prefix='analytics_export'):
    """
    Generates a multi-sheet corporate Excel workbook for executive dashboard reporting.

    :param sheets_config: List of dicts, where each dict contains:
           {
               'title': 'P&L Summary',
               'headers': ['Metric', 'Amount'],
               'rows': [['Revenue', 765.00], ...],
               'formats': ['text', 'currency']
           }
    :param filename_prefix: Base filename prefix
    :return: HttpResponse with .xlsx content type
    """
    wb = openpyxl.Workbook()
    # Remove initial default sheet
    default_sheet = wb.active

    for idx, sheet_def in enumerate(sheets_config):
        title = sheet_def.get('title', f'Sheet{idx+1}')[:31]
        ws = wb.create_sheet(title=title)

        headers = sheet_def.get('headers', [])
        rows = sheet_def.get('rows', [])
        formats = sheet_def.get('formats', ['text'] * len(headers))

        ws.append(headers)

        for row_data in rows:
            row_num = ws.max_row + 1
            for col_idx, val in enumerate(row_data, start=1):
                fmt_type = formats[col_idx - 1] if col_idx - 1 < len(formats) else 'text'
                cell = ws.cell(row=row_num, column=col_idx)
                format_cell_value(cell, val, fmt_type)

        apply_worksheet_styling(ws, freeze_header=True)

    wb.remove(default_sheet)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    date_str = timezone.now().strftime('%Y-%m-%d')
    filename = f"{filename_prefix}_{date_str}.xlsx"

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def serialize_sheets_for_preview(sheets_config):
    """
    Serializes multi-sheet export configurations into JSON-safe dictionaries
    for interactive in-browser spreadsheet previewing before download.
    """
    serialized = []
    for sheet in sheets_config:
        rows = []
        for row in sheet.get('rows', []):
            formatted_row = []
            for val in row:
                if val is None:
                    formatted_row.append('')
                elif isinstance(val, Decimal):
                    formatted_row.append(float(val))
                elif isinstance(val, (datetime, date)):
                    formatted_row.append(val.strftime('%Y-%m-%d'))
                else:
                    formatted_row.append(val)
            rows.append(formatted_row)
        serialized.append({
            'title': sheet.get('title', 'Sheet'),
            'headers': sheet.get('headers', []),
            'formats': sheet.get('formats', []),
            'rows': rows,
            'row_count': len(rows),
        })
    return serialized
